#!/usr/bin/env python3
"""
Generate Step 2 Unified Variable Table as XLSX
منصة المدقق الديناميكي الموحد V3.0
"""
import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── Sheet 1: Unified Variable Table ───
ws1 = wb.active
ws1.title = "جدول المتغيرات الموحد"
ws1.sheet_view.rightToLeft = True

header_font = Font(name='Noto Sans SC', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
input_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
lookup_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
computed_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
locked_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
output_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
data_font = Font(name='Noto Sans SC', size=10)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

type_fills = {
    'input': input_fill,
    'lookup': lookup_fill,
    'computed': computed_fill,
    'locked': locked_fill,
    'output': output_fill,
}

headers = ['#', 'الاسم الكنوني', 'الرمز', 'الفئة', 'المسار', 'الوحدة', 'الوصف العربي', 'المحرك', 'يعتمد على', 'الصيغة', 'مقفل']
ws1.append(headers)
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

variables = [
    # INPUT
    (1, 'P', 'P', 'input', 'shared', 'kg', 'وزن القنبلة', 'penetration', '', '', False),
    (2, 'lo_b', 'loб', 'input', 'shared', 'm', 'الطول الكلي للقنبلة', 'penetration', '', '', False),
    (3, 'lk', 'lk', 'input', 'shared', 'm', 'طول جسم القنبلة', 'penetration', '', '', False),
    (4, 'dk', 'dk', 'input', 'shared', 'm', 'قطر هيكل القنبلة', 'penetration', '', '', False),
    (5, 'ld_ratio', 'lᴈ/dᴈ', 'input', 'shared', '—', 'نسبة طول الجسم للقطر', 'penetration', '', '', False),
    (6, 'lhd_ratio', 'lгч/d', 'input', 'shared', '—', 'نسبة طول الرأس الحربي للقطر', 'penetration', '', '', False),
    (7, 'C', 'C', 'input', 'shared', 'kg', 'وزن الشحنة المتفجرة', 'penetration', '', '', False),
    (8, 'V', 'V', 'input', 'shared', 'm/s', 'سرعة اصطدام القنبلة', 'penetration', '', '', False),
    (9, 'alpha', 'α', 'input', 'shared', 'deg', 'زاوية الاصطدام عن الشاقول', 'penetration', '', '', False),
    (10, 'beta', 'β', 'input', 'shared', 'deg', 'زاوية ميول الجبل', 'penetration', '', '', False),
    (11, 'Z', 'Z', 'input', 'shared', 'm', 'عمق توضع المنشأة', 'penetration', '', '', False),
    # GEOMETRY INPUT
    (12, 'a_et', 'aэt', 'input', 'shared', 'm', 'ارتفاع السقف', 'blast', '', '', False),
    (13, 'bp', 'bp', 'input', 'shared', 'm', 'المجاز الطويل', 'structural', '', '', False),
    (14, 'ap', 'ap', 'input', 'shared', 'm', 'المجاز القصير', 'structural', '', '', False),
    (15, 'Lk', 'Lk', 'input', 'shared', 'm', 'طول المنشأة', 'structural', '', '', False),
    (16, 'Bk', 'Bk', 'input', 'shared', 'm', 'عرض المنشأة', 'structural', '', '', False),
    (17, 'Pk', 'Pk', 'input', 'shared', 't', 'الوزن الكلي للمنشأة', 'structural', '', '', False),
    (18, 'Hct', 'Hct', 'input', 'wall', 'm', 'سماكة الجدار الخارجي', 'structural', '', '', False),
    (19, 'Hvct', 'Hвct', 'input', 'wall', 'm', 'سماكة الجدران الداخلية', 'structural', '', '', False),
    (20, 'Hf', 'Hф', 'input', 'shared', 'm', 'سماكة الأرضية', 'structural', '', '', False),
    (21, 'Hp', 'Hп', 'input', 'roof', 'm', 'سماكة السقف', 'structural', '', '', False),
    (22, 'h_obs', 'hoбc', 'input', 'shared', 'm', 'سماكة طبقة التمويه', 'blast', '', '', False),
    (23, 'psi_p', 'ψп', 'input', 'roof', '—', 'معامل ديناميكي للسقف', 'blast', '', '', False),
    (24, 'Ea', 'Ea', 'input', 'shared', 'kg/cm2', 'معامل المرونة للحديد', 'structural', '', '', False),
    (25, 'xi', 'ξ', 'input', 'shared', '—', 'نسبة التسليح', 'structural', '', '', False),
    (26, 'rho_pc', 'ρp.c', 'input', 'roof', '—', 'معامل طبقة توزيع الضغط', 'blast', '', '', False),
    (27, 'rho_g', 'ρг', 'input', 'wall', '—', 'معامل التربة حول الجدار', 'blast', '', '', False),
    # LOOKUP
    (28, 'K1', 'K1', 'lookup', 'shared', '—', 'معامل المادة المتفجرة', 'penetration', '', '', False),
    (29, 'kpr_g', 'kпр.г', 'lookup', 'shared', '—', 'معامل اختراق التربة', 'penetration', '', '', False),
    (30, 'kpr_b', 'kпр.б', 'lookup', 'shared', '—', 'معامل اختراق البيتون', 'structural', '', '', False),
    (31, 'kpr_bt', 'kпр.бt', 'lookup', 'shared', '—', 'معامل اختراق بلاطة الحماية', 'structural', '', '', False),
    (32, 'K_kp_ct', 'Kkp.ct', 'lookup', 'wall', '—', 'معامل تدمير الجدار', 'structural', '', '', False),
    (33, 'm1', 'm1', 'lookup', 'shared', '—', 'معامل الإجهاد', 'blast', '', '', False),
    (34, 'RbH', 'RbH', 'lookup', 'shared', 'kg/cm2', 'مقاومة ضغط الخرسانة', 'structural', '', '', False),
    (35, 'RsH', 'RsH', 'lookup', 'shared', 'kg/cm2', 'إجهاد خضوع الحديد', 'structural', '', '', False),
    (36, 'gamma_b', 'γб', 'lookup', 'shared', 'kg/m3', 'الوزن الحجمي للخرسانة', 'structural', '', '', False),
    (37, 'gamma_g', 'γг', 'lookup', 'shared', 'kg/m3', 'الوزن الحجمي للتربة', 'blast', '', '', False),
    (38, 'Kpod_b', 'Kпод.b', 'lookup', 'shared', '—', 'معامل الزيادة الديناميكية للبيتون', 'structural', '', '', False),
    (39, 'Kpod_s', 'Kпод.s', 'lookup', 'shared', '—', 'معامل الزيادة الديناميكية للحديد', 'structural', '', '', False),
    (40, 'n0', 'n0', 'lookup', 'shared', '—', 'معامل الأمان', 'structural', '', '', False),
    (41, 'Kbt_bt', 'Kвt.бt', 'lookup', 'roof', '—', 'معامل التصدع لبلاطة الحماية', 'structural', '', '', False),
    (42, 'R_bar', 'R̽¯', 'lookup', 'shared', '—', 'معامل البعد المكافئ (من الجدول)', 'penetration', 'h_z_bar', '', False),
    # COMPUTED — Penetration
    (43, 'lambda1', 'λ1', 'computed', 'shared', '—', 'معامل تأثير شكل الرأس', 'penetration', 'lhd_ratio', '0.5+0.4*(lгч/d)^0.666', False),
    (44, 'lambda2', 'λ2', 'computed', 'shared', '—', 'معامل تأثير القطر', 'penetration', 'dk', '2.8*(d)^0.333-1.3*(d)^0.5', False),
    (45, 'tsu', 'ц', 'computed', 'shared', 'm', 'إزاحة مركز الانفجار', 'penetration', 'lk,alpha', '0.5*lk*cos(α)', False),
    (46, 'H_total', 'H', 'computed', 'shared', 'm', 'الارتفاع من سقف النفق حتى سطح الجبل', 'penetration', 'R_actual,h_pr,tsu', 'R̽+hпр-ц', False),
    (47, 'h_z', 'hз', 'computed', 'shared', 'm', 'عمق ما دون الحفرة', 'penetration', 'h_pr,tsu', 'hпр-ц', False),
    (48, 'h_z_bar', 'hз¯', 'computed', 'shared', '—', 'العمق المكافئ المختزل', 'penetration', 'h_z,C_ef', 'hз/(Cэф)^0.333', False),
    (49, 'L_tunnel', 'L', 'computed', 'shared', 'm', 'طول المدخل ضمن الحمل الديناميكي', 'penetration', 'H_total,beta', 'H/tan(β)', False),
    (50, 'Y_diff', 'Y', 'computed', 'shared', 'm', 'الفرق بين عمق التوضع والاختراق', 'penetration', 'Z,h_pr', 'Z-hпр', False),
    (51, 'hb_destruction', 'hb.dest', 'computed', 'shared', 'm', 'سماكة البيتون للتدمير', 'penetration', 'Zp,Y_diff', '(Zp-Y)*(Kp.b/Kp.r)', False),
    (52, 'hb_cracking', 'hb.crack', 'computed', 'shared', 'm', 'سماكة البيتون للتصدع', 'penetration', 'Zot_rock,Y_diff', '(Zot-Y)*(Kot.b/Kot.r)', False),
    (53, 'Zot_rock', 'Zot.rock', 'computed', 'shared', 'm', 'عمق التصدع الصخري', 'penetration', 'C_ef', '1.5*kot.г*(Cэф)^0.333', False),
    # LOCKED — Penetration → Blast
    (54, 'C_ef', 'Cэф', 'locked', 'shared', 'kg', 'الشحنة الفعالة', 'penetration', 'K1,C', 'K1*0.95*C', True),
    (55, 'h_pr', 'hпр', 'locked', 'shared', 'm', 'عمق اختراق القنبلة', 'penetration', 'λ1,λ2,kпр.г,P,d,V,α', 'λ1*λ2*kпр.г*P/d^2*V*cosα', True),
    (56, 'R_actual', 'R̽', 'locked', 'shared', 'm', 'البعد الفعلي لمركز الانفجار', 'penetration', 'R_bar,C_ef', 'R̽¯*(Cэф)^0.333', True),
    (57, 'Zp', 'Zp', 'locked', 'shared', 'm', 'عمق التربة المكافئ', 'penetration', 'kpr_g,C_ef', '1.5*kпр.г*(Cэф)^0.333', True),
    # COMPUTED — Blast Roof
    (58, 'ht', 'ht', 'computed', 'roof', 'm', 'سماكة بلاطة الحماية', 'blast', 'h_pr,a_et,Z', '', False),
    (59, 'R0_roof', 'R0', 'computed', 'roof', 'm', 'بعد الانفجار عن السقف', 'blast', 'R_actual,h_pr,Z', '', False),
    (60, 'Hpc', 'Hpc', 'computed', 'roof', 'm', 'سماكة طبقة توزيع الضغط', 'blast', 'ht,Z,R_actual', '', False),
    (61, 'R_ekv_roof', 'Rэкв', 'computed', 'roof', 'm', 'البعد المكافئ — السقف', 'blast', 'R0_roof,Hpc', '', False),
    (62, 'R_star_roof', 'R*', 'computed', 'roof', 'm', 'البعد الفعال — السقف', 'blast', 'R_ekv_roof,ht', '', False),
    (63, 'max_bv_roof', 'maxбв', 'computed', 'roof', 'kg/cm2', 'الإجهاد الأقصى في التربة — السقف', 'blast', 'R_star_roof,C_ef,m1', '', False),
    (64, 'tau_roof', 'τ', 'computed', 'roof', 's', 'مدة تأثير الحمل — السقف', 'blast', 'R_ekv_roof,a0cp,a1cp', '', False),
    (65, 'tau_n_roof', 'τн', 'computed', 'roof', 's', 'زمن الصعود — السقف', 'blast', 'R_ekv_roof,a0cp', '', False),
    (66, 'Zp_roof', 'Zп', 'computed', 'roof', 'm', 'البعد المختزل — السقف', 'blast', 'R_ekv_roof,C_ef', '', False),
    (67, 'a0cp_roof', 'a0cp', 'computed', 'roof', 'm/s', 'سرعة الموجة P — السقف', 'blast', 'rho_pc', '', False),
    (68, 'a1cp_roof', 'a1cp', 'computed', 'roof', 'm/s', 'سرعة الموجة S — السقف', 'blast', 'a0cp_roof', '', False),
    # LOCKED — Blast Roof
    (69, 'Bt', 'Bt', 'locked', 'roof', 'm', 'بروز بلاطة الحماية', 'blast', 'ht,bp,ap', '', True),
    (70, 'tau_ef_roof', 'τэф', 'locked', 'roof', 's', 'الزمن الفعال — السقف', 'blast', 'tau_roof,tau_n_roof', 'τ-maxбв/τн', True),
    # COMPUTED — Blast Wall
    (71, 'Z_wall', 'Zc', 'computed', 'wall', 'm', 'البعد حتى الجدار', 'blast', 'Zp_roof,Z', '', False),
    (72, 'h_b', 'hб', 'computed', 'wall', 'm', 'سماكة البيتون — الجدار', 'blast', 'Zp,Y_diff,K_kp_ct', '', False),
    (73, 'tau_theta', 'τθ', 'computed', 'wall', '—', 'معامل زاوي — الجدار', 'blast', '', '', False),
    (74, 'h_exposure_wall', 'h_exp.w', 'computed', 'wall', 'm', 'ارتفاع التعرض — الجدار', 'blast', '', '', False),
    (75, 'R_ekv_wall', 'Rэкв.wall', 'computed', 'wall', 'm', 'البعد المكافئ — الجدار', 'blast', '', '', False),
    (76, 'R_star_wall', 'R*.wall', 'computed', 'wall', 'm', 'البعد الفعال — الجدار', 'blast', '', '', False),
    (77, 'max_bv_wall', 'maxбв.wall', 'computed', 'wall', 'kg/cm2', 'الإجهاد الأقصى في التربة — الجدار', 'blast', '', '', False),
    (78, 'tau_wall', 'τ.wall', 'computed', 'wall', 's', 'مدة تأثير الحمل — الجدار', 'blast', '', '', False),
    (79, 'tau_n_wall', 'τн.wall', 'computed', 'wall', 's', 'زمن الصعود — الجدار', 'blast', '', '', False),
    (80, 'Zc_wall', 'Zc.wall', 'computed', 'wall', 'm', 'البعد المختزل — الجدار', 'blast', '', '', False),
    (81, 'a0cp_wall', 'a0cp.wall', 'computed', 'wall', 'm/s', 'سرعة الموجة P — الجدار', 'blast', '', '', False),
    (82, 'a1cp_wall', 'a1cp.wall', 'computed', 'wall', 'm/s', 'سرعة الموجة S — الجدار', 'blast', '', '', False),
    # LOCKED — Blast Wall
    (83, 'tau_ef_wall', 'τэф.wall', 'locked', 'wall', 's', 'الزمن الفعال — الجدار', 'blast', 'tau_wall,tau_n_wall', 'τ.wall-maxбв.wall/τн.wall', True),
    # LOCKED — Governing
    (84, 'omega', 'ω', 'locked', 'shared', 's-1', 'التردد الدائري للحمل الديناميكي', 'blast', 'tau_ef_roof,tau_n_roof', '', True),
    (85, 'C_dyn_roof', 'Cд.roof', 'computed', 'roof', '—', 'معامل الديناميكية — السقف', 'blast', 'omega,tau_ef_roof', '', False),
    (86, 'C_dyn_wall', 'Cд.wall', 'computed', 'wall', '—', 'معامل الديناميكية — الجدار', 'blast', 'omega,tau_ef_wall', '', False),
    (87, 'mu_roof', 'μ.roof', 'computed', 'roof', '—', 'معامل المطاوعة — السقف', 'structural', '', '', False),
    (88, 'mu_wall', 'μ.wall', 'computed', 'wall', '—', 'معامل المطاوعة — الجدار', 'structural', '', '', False),
    # LOCKED — Pressures Roof
    (89, 'Pmax_roof', 'Pmax.roof', 'locked', 'roof', 'kg/cm2', 'الضغط الأقصى — السقف', 'blast', 'max_bv_roof,C_ef,R_ekv_roof', '', True),
    (90, 'P_ekv_roof', 'Pэкв.roof', 'locked', 'roof', 'kg/cm2', 'الضغط المكافئ — السقف', 'blast', 'Pmax_roof,C_dyn_roof', '', True),
    (91, 'Pp_roof', 'Pp.roof', 'locked', 'roof', 'kg/cm2', 'الحمولة الحسابية على السقف', 'structural', 'P_ekv_roof,Pct_roof', '', True),
    (92, 'Pct_roof', 'Pct.roof', 'computed', 'roof', 'kg/cm2', 'حمولة الجدار المؤثرة على السقف', 'structural', '', '', False),
    # LOCKED — Pressures Wall
    (93, 'Pmax_wall', 'Pmax.wall', 'locked', 'wall', 'kg/cm2', 'الضغط الأقصى — الجدار', 'blast', 'max_bv_wall,C_ef,R_ekv_wall', '', True),
    (94, 'P_ekv_wall', 'Pэкв.wall', 'locked', 'wall', 'kg/cm2', 'الضغط المكافئ — الجدار', 'blast', 'Pmax_wall,C_dyn_wall', '', True),
    (95, 'Pp_wall', 'Pp.wall', 'locked', 'wall', 'kg/cm2', 'الحمولة الحسابية على الجدار', 'structural', 'P_ekv_wall,Pct_wall', '', True),
    (96, 'Pct_wall', 'Pct.wall', 'computed', 'wall', 'kg/cm2', 'حمولة السقف المؤثرة على الجدار', 'structural', '', '', False),
    # OUTPUT
    (97, 'Hc_output', 'Hc', 'output', 'wall', 'm', 'سماكة الجدار المطلوبة', 'structural', 'Pp_wall,RbH,n0', '', False),
    (98, 'Hp_output', 'Hп.out', 'output', 'roof', 'm', 'سماكة السقف المطلوبة', 'structural', 'Pp_roof,RbH,n0', '', False),
    (99, 'Hf_output', 'Hф.out', 'output', 'shared', 'm', 'سماكة الأرضية المطلوبة', 'structural', 'Pp_roof,Pp_wall', '', False),
    (100, 'status', 'status', 'output', 'shared', '—', 'حالة التحقق الإنشائي', 'structural', 'Hc_output,Hp_output,Hf_output', '', False),
]

for row_data in variables:
    idx, name, symbol, vtype, path, unit, desc, engine, deps, formula, locked = row_data
    row = [idx, name, symbol, vtype, path, unit, desc, engine, deps, formula, 'نعم' if locked else 'لا']
    ws1.append(row)
    fill = type_fills.get(vtype, PatternFill())
    for col_idx, val in enumerate(row, 1):
        cell = ws1.cell(row=ws1.max_row, column=col_idx, value=val)
        cell.font = data_font
        cell.fill = fill
        cell.border = border
        cell.alignment = center if col_idx in [1,3,4,5,6,8,11] else right_align

# Column widths
widths = [5, 18, 14, 10, 10, 10, 35, 14, 25, 30, 8]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ─── Sheet 2: Reference Values (Step 3) ───
ws2 = wb.create_sheet("القيم المرجعية BMK-02")
ws2.sheet_view.rightToLeft = True

ref_data = [
    ('المدخلات الأساسية', 'input', [
        ('P', 'P', 'kg', 441),
        ('lo_b', 'loб', 'm', 3.01),
        ('lk', 'lk', 'm', 1.92),
        ('dk', 'dk', 'm', 0.36),
        ('ld_ratio', 'lᴈ/dᴈ', '—', 5.3),
        ('lhd_ratio', 'lгч/d', '—', 2),
        ('C', 'C', 'kg', 215),
        ('V', 'V', 'm/s', 350),
        ('alpha', 'α', 'deg', 20),
        ('beta', 'β', 'deg', 22),
        ('Z', 'Z', 'm', 3.7),
    ]),
    ('معاملات التربة والمواد', 'lookup', [
        ('K1', 'K1', '—', 1.639),
        ('kpr_g', 'kпр.г', '—', 1.8e-6),
        ('kpr_b', 'kпр.б', '—', 8e-7),
        ('kpr_bt', 'kпр.бt', '—', 8e-7),
        ('K_kp_ct', 'Kkp.ct', '—', 1.1),
        ('m1', 'm1', '—', 1.65),
        ('RbH', 'RbH', 'kg/cm2', 200),
        ('RsH', 'RsH', 'kg/cm2', 3000),
        ('gamma_b', 'γб', 'kg/m3', 2500),
        ('gamma_g', 'γг', 'kg/m3', 1700),
        ('Kpod_b', 'Kпод.b', '—', 1.18),
        ('Kpod_s', 'Kпод.s', '—', 1.25),
        ('n0', 'n0', '—', 1.25),
        ('Kbt_bt', 'Kвt.бt', '—', 0.13),
    ]),
    ('أبعاد المنشأة', 'input', [
        ('a_et', 'aэt', 'm', 3),
        ('bp', 'bp', 'm', 5),
        ('ap', 'ap', 'm', 4),
        ('Lk', 'Lk', 'm', 50),
        ('Bk', 'Bk', 'm', 20),
        ('Pk', 'Pk', 't', 492.25),
        ('Hct', 'Hct', 'm', 0.5),
        ('Hvct', 'Hвct', 'm', 0.3),
        ('Hf', 'Hф', 'm', 0.45),
        ('Hp', 'Hп', 'm', 0.75),
        ('h_obs', 'hoбc', 'm', 0.5),
        ('psi_p', 'ψп', '—', 0.012),
        ('Ea', 'Ea', 'kg/cm2', 2100000),
        ('xi', 'ξ', '—', 0.55),
        ('rho_pc', 'ρp.c', '—', 150),
        ('rho_g', 'ρг', '—', 180),
    ]),
    ('نتائج الاختراق (مقفلة)', 'locked', [
        ('C_ef', 'Cэф', 'kg', 334.76575),
        ('lambda1', 'λ1', '—', 1.134667074552914),
        ('lambda2', 'λ2', '—', 1.21253869486675),
        ('h_pr', 'hпр', 'm', 3.6495332546231958),
        ('R_actual', 'R̽', 'm', 7.6230969724513375),
        ('h_z', 'hз', 'm', 2.6895332546232003),
        ('h_z_bar', 'hз¯', '—', 0.388095099770739),
        ('Zp', 'Zp', 'm', 5.8212740516901125),
        ('L_tunnel', 'L', 'm', 25.53896746890041),
        ('Y_diff', 'Y', 'm', 0.05046674537679996),
        ('hb_destruction', 'hb.dest', 'm', 3.297604175036178),
        ('hb_cracking', 'hb.crack', 'm', 4.752375128334365),
        ('Zot_rock', 'Zot.rock', 'm', 8.212154465777123),
    ]),
    ('نتائج السقف (مقفلة)', 'locked', [
        ('ht', 'ht', 'm', 1.0721670569007022),
        ('Bt', 'Bt', 'm', 8.05201583975881),
        ('R0_roof', 'R0', 'm', 7.043741602545037),
        ('Hpc', 'Hpc', 'm', 3.8320486334489554),
        ('R_ekv_roof', 'Rэкв', 'm', 6.1162229173010605),
        ('R_star_roof', 'R*', 'm', 2.4255308548708796),
        ('max_bv_roof', 'maxбв', 'kg/cm2', 2.880259056606225),
        ('tau_roof', 'τ', 's', 0.2649955476788649),
        ('tau_ef_roof', 'τэф', 's', 0.2377897177708615),
        ('tau_n_roof', 'τн', 's', 0.036467651189518974),
        ('Zp_roof', 'Zп', 'm', 5.40421569034966),
        ('a0cp_roof', 'a0cp', 'm/s', 152.5),
        ('a1cp_roof', 'a1cp', 'm/s', 60.83333333333333),
    ]),
    ('نتائج الجدار (مقفلة)', 'locked', [
        ('Z_wall', 'Zc', 'm', 7.50421569034966),
        ('h_b', 'hб', 'm', 3.441740788929493),
        ('tau_theta', 'τθ', '—', 0.5767495644576623),
        ('h_exposure_wall', 'h_exp.w', 'm', 0.4966373747446921),
        ('R_ekv_wall', 'Rэкв.wall', 'm', 7.04374160254504),
        ('R_star_wall', 'R*.wall', 'm', 6.237079341096549),
        ('max_bv_wall', 'maxбв.wall', 'kg/cm2', 3.142823347197694),
        ('tau_wall', 'τ.wall', 's', 0.06848701632826656),
        ('tau_ef_wall', 'τэф.wall', 's', 0.06085190940339927),
        ('tau_n_wall', 'τн.wall', 's', 0.011111111111111112),
        ('Zc_wall', 'Zc.wall', 'm', 7.50421569034966),
        ('a0cp_wall', 'a0cp.wall', 'm/s', 540),
        ('a1cp_wall', 'a1cp.wall', 'm/s', 270),
    ]),
    ('القيم الحاكمة (مقفلة)', 'locked', [
        ('omega', 'ω', 's-1', 561.6673670487412),
        ('C_dyn_roof', 'Cд.roof', '—', 46.81109581088254),
        ('C_dyn_wall', 'Cд.wall', '—', 72.08111111111113),
        ('mu_roof', 'μ.roof', '—', 0.8861874999999999),
        ('mu_wall', 'μ.wall', '—', 0.912333333333333),
        ('Pmax_roof', 'Pmax.roof', 'kg/cm2', 4.60841449056996),
        ('P_ekv_roof', 'Pэкв.roof', 'kg/cm2', 3.8157671981919274),
        ('Pp_roof', 'Pp.roof', 'kg/cm2', 4.92111625743445),
        ('Pct_roof', 'Pct.roof', 'kg/cm2', 1.1053490592425188),
        ('Pmax_wall', 'Pmax.wall', 'kg/cm2', 6.285646694395388),
        ('P_ekv_wall', 'Pэкв.wall', 'kg/cm2', 3.082860450496599),
        ('Pp_wall', 'Pp.wall', 'kg/cm2', 3.784504617544293),
        ('Pct_wall', 'Pct.wall', 'kg/cm2', 0.7016441670476933),
    ]),
]

ref_headers = ['القسم', 'الفئة', 'الاسم', 'الرمز', 'الوحدة', 'القيمة المرجعية']
ws2.append(ref_headers)
for col_idx, h in enumerate(ref_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

for section_name, section_type, items in ref_data:
    for name, symbol, unit, value in items:
        row = [section_name, section_type, name, symbol, unit, value]
        ws2.append(row)
        fill = type_fills.get(section_type, PatternFill())
        for col_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=ws2.max_row, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.border = border
            if col_idx == 6 and isinstance(val, float):
                cell.number_format = '0.000000000000'
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = center

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 28

# ─── Sheet 3: Step 4 Intermediate Values ───
ws3 = wb.create_sheet("القيم الوسيطة للخطوة 4")
ws3.sheet_view.rightToLeft = True

s4_headers = ['الاسم', 'الرمز', 'الوحدة', 'القيمة', 'الوصف']
ws3.append(s4_headers)
for col_idx, h in enumerate(s4_headers, 1):
    cell = ws3.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='8E24AA', end_color='8E24AA', fill_type='solid')
    cell.alignment = center
    cell.border = border

step4_data = [
    ('h_pr', 'hпр', 'm', 3.6495332546231958, 'عمق اختراق القنبلة'),
    ('Zp', 'Zp', 'm', 5.8212740516901125, 'عمق التربة المكافئ'),
    ('R_actual', 'R̽', 'm', 7.6230969724513375, 'البعد الفعلي لمركز الانفجار'),
    ('C_ef', 'Cэф', 'kg', 334.76575, 'الشحنة الفعالة'),
    ('tau_ef_roof', 'τэф', 's', 0.2377897177708615, 'الزمن الفعال — السقف'),
    ('tau_ef_wall', 'τэф.wall', 's', 0.06085190940339927, 'الزمن الفعال — الجدار'),
    ('omega', 'ω', 's-1', 561.6673670487412, 'التردد الدائري للحمل'),
    ('Pmax_roof', 'Pmax.roof', 'kg/cm2', 4.60841449056996, 'الضغط الأقصى — السقف'),
    ('Pmax_wall', 'Pmax.wall', 'kg/cm2', 6.285646694395388, 'الضغط الأقصى — الجدار'),
    ('P_ekv_roof', 'Pэкв.roof', 'kg/cm2', 3.8157671981919274, 'الضغط المكافئ — السقف'),
    ('P_ekv_wall', 'Pэкв.wall', 'kg/cm2', 3.082860450496599, 'الضغط المكافئ — الجدار'),
    ('Pp_roof', 'Pp.roof', 'kg/cm2', 4.92111625743445, 'الحمولة الحسابية — السقف'),
    ('Pp_wall', 'Pp.wall', 'kg/cm2', 3.784504617544293, 'الحمولة الحسابية — الجدار'),
    ('Hp', 'Hп', 'm', 0.75, 'سماكة السقف'),
    ('Hct', 'Hc', 'm', 0.5, 'سماكة الجدار'),
    ('Hf', 'Hф', 'm', 0.45, 'سماكة الأرضية'),
    ('Bt', 'Bt', 'm', 8.05201583975881, 'بروز بلاطة الحماية'),
]

locked_fill_purple = PatternFill(start_color='E1BEE7', end_color='E1BEE7', fill_type='solid')
for name, symbol, unit, value, desc in step4_data:
    row = [name, symbol, unit, value, desc]
    ws3.append(row)
    for col_idx, val in enumerate(row, 1):
        cell = ws3.cell(row=ws3.max_row, column=col_idx, value=val)
        cell.font = data_font
        cell.fill = locked_fill_purple
        cell.border = border
        if col_idx == 4 and isinstance(val, float):
            cell.number_format = '0.000000000000'
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.alignment = center

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 28
ws3.column_dimensions['E'].width = 35

# ─── Sheet 4: Locked Registry ───
ws4 = wb.create_sheet("سجل القيم المقفلة")
ws4.sheet_view.rightToLeft = True

l4_headers = ['الاسم', 'الرمز', 'ينتجه', 'يستهلكه', 'المسار', 'القراءة فقط']
ws4.append(l4_headers)
for col_idx, h in enumerate(l4_headers, 1):
    cell = ws4.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
    cell.alignment = center
    cell.border = border

locked_registry = [
    ('C_ef', 'Cэф', 'penetration', 'blast', 'shared', 'نعم'),
    ('h_pr', 'hпр', 'penetration', 'blast', 'shared', 'نعم'),
    ('R_actual', 'R̽', 'penetration', 'blast', 'shared', 'نعم'),
    ('Zp', 'Zp', 'penetration', 'blast, structural', 'shared', 'نعم'),
    ('Bt', 'Bt', 'blast', 'structural', 'roof', 'نعم'),
    ('tau_ef_roof', 'τэф', 'blast', 'blast, structural', 'roof', 'نعم'),
    ('tau_ef_wall', 'τэф.wall', 'blast', 'blast, structural', 'wall', 'نعم'),
    ('omega', 'ω', 'blast', 'blast, structural', 'shared', 'نعم'),
    ('Pmax_roof', 'Pmax.roof', 'blast', 'structural', 'roof', 'نعم'),
    ('P_ekv_roof', 'Pэкв.roof', 'blast', 'structural', 'roof', 'نعم'),
    ('Pp_roof', 'Pp.roof', 'structural', 'structural', 'roof', 'نعم'),
    ('Pmax_wall', 'Pmax.wall', 'blast', 'structural', 'wall', 'نعم'),
    ('P_ekv_wall', 'Pэкв.wall', 'blast', 'structural', 'wall', 'نعم'),
    ('Pp_wall', 'Pp.wall', 'structural', 'structural', 'wall', 'نعم'),
]

for row_data in locked_registry:
    ws4.append(row_data)
    for col_idx, val in enumerate(row_data, 1):
        cell = ws4.cell(row=ws4.max_row, column=col_idx, value=val)
        cell.font = data_font
        cell.fill = locked_fill
        cell.border = border
        cell.alignment = center

ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 14
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 24
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 14

# ─── Save ───
output_path = '/home/z/my-project/download/step2-unified-variable-model.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
