#!/usr/bin/env python3
"""Generate Steps 2-8 Unified Variable Model as XLSX"""
import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
hf = Font(name='Noto Sans SC', size=11, bold=True, color='FFFFFF')
df = Font(name='Noto Sans SC', size=10)
c = Alignment(horizontal='center', vertical='center', wrap_text=True)
r = Alignment(horizontal='right', vertical='center', wrap_text=True)
b = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

fills = {
    'input': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    'lookup': PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid'),
    'computed': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    'locked': PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid'),
    'output': PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid'),
}
hf_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
step_fill = PatternFill(start_color='8E24AA', end_color='8E24AA', fill_type='solid')
locked_red = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
purple_fill = PatternFill(start_color='E1BEE7', end_color='E1BEE7', fill_type='solid')

def style_header(ws, fill=hf_fill):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = hf
        cell.fill = fill
        cell.alignment = c
        cell.border = b

def add_data_row(ws, row_data, fill_map=None, key_col=3):
    ws.append(row_data)
    fill = fill_map.get(row_data[key_col], PatternFill()) if fill_map else PatternFill()
    for col_idx in range(1, len(row_data) + 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = df
        cell.fill = fill
        cell.border = b
        cell.alignment = c

# ─── Sheet 1: Final Locked Results ───
ws1 = wb.active
ws1.title = "النتائج النهائية المقفلة"
ws1.sheet_view.rightToLeft = True
ws1.append(['الرمز', 'الوحدة', 'القيمة', 'الوصف', 'الخطوة'])
style_header(ws1, step_fill)

final = [
    ('Hп', 'cm', 70.4594848625, 'سماكة السقف النهائية', '7'),
    ('Hc', 'cm', 49.8223795452, 'سماكة الجدار النهائية', '8'),
    ('Hф', 'cm', 42.3490226134, 'سماكة الأرضية النهائية', '8'),
    ('Hвc', 'cm', 30, 'سماكة الجدران الداخلية', '8'),
    ('ht', 'cm', 107.2167056901, 'سماكة بلاطة الحماية', '4'),
    ('Bt', 'm', 8.0520158398, 'بروز البلاطة', '4'),
    ('Hp.c', 'm', 3.8320486334, 'طبقة توزيع الضغط', '4'),
    ('Pp(roof)', 'kg/cm2', 4.9211162574, 'الحمل الحسابي سقف', '5'),
    ('Pp(wall)', 'kg/cm2', 3.7845046175, 'الحمل الحسابي جدار', '8'),
    ('Mp(roof)', 'kg.cm', 20000000, 'العزم سقف', '7'),
    ('Mp(wall)', 'kg.cm', 10000000, 'العزم جدار', '8'),
]
for d in final:
    ws1.append(d)
    for col_idx in range(1, len(d)+1):
        cell = ws1.cell(row=ws1.max_row, column=col_idx)
        cell.font = df; cell.fill = purple_fill; cell.border = b; cell.alignment = c
        if col_idx == 3 and isinstance(d[2], float):
            cell.number_format = '0.0000000000'

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 24
ws1.column_dimensions['D'].width = 30
ws1.column_dimensions['E'].width = 8

# ─── Sheet 2: Step 5 Roof ───
ws2 = wb.create_sheet("الخطوة 5 — السقف")
ws2.sheet_view.rightToLeft = True
ws2.append(['الرمز', 'الوحدة', 'القيمة', 'الوصف', 'القسم'])
style_header(ws2)

roof_data = [
    ('h̄', '—', 0.1180386444, 'الارتفاع المختزل', '5.1'),
    ('R̄̽', '—', 0.35, 'معامل البعد B-1', '5.1'),
    ('Rэкв', 'm', 6.1162229173, 'البعد المكافئ', '5.1'),
    ('R*', 'm', 2.4255308549, 'البعد الفعال', '5.1'),
    ('maxбв', 'kg/cm2', 2.8802590566, 'الإجهاد الأقصى', '5.1'),
    ('τ', 's', 0.2649955477, 'زمن الحمل', '5.3'),
    ('τэф', 's', 0.2377897178, 'الزمن الفعال', '5.4'),
    ('τн', 's', 0.0364676512, 'زمن الصعود', '5.5'),
    ('a0cp', 'm/s', 152.5, 'سرعة الموجة P', '5.5'),
    ('a1cp', 'm/s', 60.8333333333, 'سرعة الموجة S', '5.5'),
    ('ω', 's-1', 561.6673670487, 'التردد الدائري', '5.6'),
    ('Cд', '—', 46.8110958109, 'معامل الديناميكية', '5.6'),
    ('μ', '—', 0.8861875, 'معامل المطاوعة', '5.6'),
    ('η', '—', 1.25, 'معامل η', '5.6'),
    ('Rsd', 'kg/cm2', 3937.5, 'مقاومة الحديد الديناميكية', '5.6'),
    ('Rbd', 'kg/cm2', 236, 'مقاومة البيتون الديناميكية', '5.6'),
    ('λ', '—', 0.124184033, 'معامل λ', '5.7'),
    ('Kп', '—', 0.8, 'معامل Kп', '5.7'),
    ('Pmax', 'kg/cm2', 4.6084144906, 'الحمولة العظمى', '5.7'),
    ('Kд', '—', 0.92, 'معامل Kд', '5.8'),
    ('kψ', '—', 0.9, 'معامل kψ', '5.8'),
    ('Pэкв', 'kg/cm2', 3.8157671982, 'الحمولة المكافئة', '5.8'),
    ('Pct', 'kg/cm2', 1.1053490592, 'حمولة الجدار على السقف', '5.9'),
    ('Pp', 'kg/cm2', 4.9211162574, 'الحمولة الحسابية', '5.9'),
]
for d in roof_data:
    ws2.append(d)
    for col_idx in range(1, len(d)+1):
        cell = ws2.cell(row=ws2.max_row, column=col_idx)
        cell.font = df; cell.fill = fills.get('computed', PatternFill()); cell.border = b; cell.alignment = c
        if col_idx == 3 and isinstance(d[2], float):
            cell.number_format = '0.0000000000'

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 24
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 8

# ─── Sheet 3: Step 5/8 Wall ───
ws3 = wb.create_sheet("الخطوة 8 — الجدار")
ws3.sheet_view.rightToLeft = True
ws3.append(['الرمز', 'الوحدة', 'القيمة', 'الوصف', 'القسم'])
style_header(ws3)

wall_data = [
    ('τθ', '—', 0.5767495645, 'معامل زاوي', '8.1'),
    ('Z', 'm', 7.5042156903, 'البعد حتى الجدار', '8.1'),
    ('hб', 'm', 3.4417407889, 'سماكة البيتون', '8.1'),
    ('h̄', '—', 0.4966373747, 'الارتفاع المختزل', '8.2'),
    ('R̄̽', '—', 0.9, 'معامل البعد B-1', '8.2'),
    ('Rэкв', 'm', 7.0437416025, 'البعد المكافئ', '8.2'),
    ('R*', 'm', 6.2370793411, 'البعد الفعال', '8.2'),
    ('maxбв', 'kg/cm2', 3.1428233472, 'الإجهاد الأقصى', '8.2'),
    ('τ', 's', 0.0684870163, 'زمن الحمل', '8.3'),
    ('τэф', 's', 0.0608519094, 'الزمن الفعال', '8.4'),
    ('τн', 's', 0.0111111111, 'زمن الصعود', '8.5'),
    ('a0cp', 'm/s', 540, 'سرعة الموجة P', '8.5'),
    ('a1cp', 'm/s', 270, 'سرعة الموجة S', '8.5'),
    ('ω', 's-1', 1024.0477954056, 'التردد الدائري', '8.6'),
    ('Cд', '—', 72.0811111111, 'معامل الديناميكية', '8.6'),
    ('μ', '—', 0.9123333333, 'معامل المطاوعة', '8.6'),
    ('η', '—', 1.6666666667, 'معامل η', '8.6'),
    ('Rsd', 'kg/cm2', 3937.5, 'مقاومة الحديد', '8.6'),
    ('Rbd', 'kg/cm2', 236, 'مقاومة البيتون', '8.6'),
    ('λ', '—', 3.1449305556, 'معامل λ', '8.7'),
    ('Kп', '—', 1, 'معامل Kп', '8.7'),
    ('Pmax', 'kg/cm2', 6.2856466944, 'الحمولة العظمى', '8.7'),
    ('Kд', '—', 1, 'معامل Kд', '8.8'),
    ('kψ', '—', 0.85, 'معامل kψ', '8.8'),
    ('Pэкв', 'kg/cm2', 3.0828604505, 'الحمولة المكافئة', '8.8'),
    ('Pct', 'kg/cm2', 0.7016441670, 'حمولة السقف على الجدار', '8.9'),
    ('Pp', 'kg/cm2', 3.7845046175, 'الحمولة الحسابية', '8.9'),
    ('Hc', 'cm', 49.8223795452, 'سماكة الجدار النهائية', '8'),
    ('Hф', 'cm', 42.3490226134, 'سماكة الأرضية النهائية', '8'),
    ('Hвc', 'cm', 30, 'سماكة الجدران الداخلية', '8'),
]
for d in wall_data:
    ws3.append(d)
    for col_idx in range(1, len(d)+1):
        cell = ws3.cell(row=ws3.max_row, column=col_idx)
        cell.font = df; cell.fill = fills.get('computed', PatternFill()); cell.border = b; cell.alignment = c
        if col_idx == 3 and isinstance(d[2], float):
            cell.number_format = '0.0000000000'

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 24
ws3.column_dimensions['D'].width = 30
ws3.column_dimensions['E'].width = 8

# ─── Sheet 4: Step 6 Table Lookups ───
ws4 = wb.create_sheet("الخطوة 6 — جداول B")
ws4.sheet_view.rightToLeft = True
ws4.append(['الجدول', 'الرمز', 'السقف', 'الجدار', 'الوحدة', 'الوصف'])
style_header(ws4, locked_red)

tables = [
    ('B-1', 'R̄̽', 0.35, 0.9, '—', 'معامل البعد المكافئ'),
    ('B-2', 'μ', 0.025, 0.009, '—', 'معامل ديناميكي'),
    ('B-2', 'η', 0.015, 0.001, '—', 'معامل η'),
    ('B-2', 'Kt', 1, 1.1, '—', 'معامل Kt'),
    ('B-3', 'a0z', 180, 580, 'm/s', 'سرعة الموجة P'),
    ('B-3', 'a1z', 80, 290, 'm/s', 'سرعة الموجة S'),
    ('B-4', 'Kпод', 1.25, 1.18, '—', 'معامل الزيادة'),
    ('B-5', 'Kп', 0.8, 1, '—', 'معامل الحمولة'),
    ('B-6', 'Kд', 0.92, 1, '—', 'معامل الديناميكية'),
]
for d in tables:
    ws4.append(d)
    for col_idx in range(1, len(d)+1):
        cell = ws4.cell(row=ws4.max_row, column=col_idx)
        cell.font = df; cell.fill = fills.get('lookup', PatternFill()); cell.border = b; cell.alignment = c

ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 10
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 16
ws4.column_dimensions['E'].width = 10
ws4.column_dimensions['F'].width = 30

# ─── Sheet 5: Step 7 Ceiling ───
ws5 = wb.create_sheet("الخطوة 7 — السقف النهائي")
ws5.sheet_view.rightToLeft = True
ws5.append(['الرمز', 'الوحدة', 'القيمة', 'الوصف'])
style_header(ws5, step_fill)

step7 = [
    ('Mp', 'kg.cm', 20000000, 'العزم الأكبر'),
    ('μ', '—', 0.8861875, 'معامل المطاوعة الإنشائي'),
    ('Rsd', 'kg/cm2', 3937.5, 'مقاومة الحديد الديناميكية'),
    ('h0', 'cm', 67.1042712976, 'العمق الفعال'),
    ('Hп', 'cm', 70.4594848625, 'السماكة النهائية = h0 × 1.05'),
]
for d in step7:
    ws5.append(d)
    for col_idx in range(1, len(d)+1):
        cell = ws5.cell(row=ws5.max_row, column=col_idx)
        cell.font = df; cell.fill = fills.get('output', PatternFill()); cell.border = b; cell.alignment = c
        if col_idx == 3 and isinstance(d[2], float):
            cell.number_format = '0.0000000000'

ws5.column_dimensions['A'].width = 10
ws5.column_dimensions['B'].width = 10
ws5.column_dimensions['C'].width = 24
ws5.column_dimensions['D'].width = 35

# ─── Sheet 6: Locked Registry ───
ws6 = wb.create_sheet("سجل القيم المقفلة")
ws6.sheet_view.rightToLeft = True
ws6.append(['الاسم', 'الرمز', 'ينتجه', 'يستهلكه', 'المسار'])
style_header(ws6, locked_red)

reg = [
    ('C_ef','Cэф','penetration','blast','shared'),
    ('h_pr','hпр','penetration','blast','shared'),
    ('R_actual','R̽','penetration','blast','shared'),
    ('Zp','Zp','penetration','blast,structural','shared'),
    ('ht','ht','blast','structural','roof'),
    ('Bt','Bt','blast','structural','roof'),
    ('Hpc','Hp.c','blast','structural','roof'),
    ('R_ekv_roof','Rэкв','blast','structural','roof'),
    ('tau_ef_roof','τэф','blast','structural','roof'),
    ('tau_ef_wall','τэф.wall','blast','structural','wall'),
    ('omega_roof','ω.roof','blast','structural','roof'),
    ('omega_wall','ω.wall','blast','structural','wall'),
    ('Pmax_roof','Pmax.roof','blast','structural','roof'),
    ('P_ekv_roof','Pэкв.roof','blast','structural','roof'),
    ('Pp_roof','Pp.roof','structural','structural','roof'),
    ('Pmax_wall','Pmax.wall','blast','structural','wall'),
    ('P_ekv_wall','Pэкв.wall','blast','structural','wall'),
    ('Pp_wall','Pp.wall','structural','structural','wall'),
    ('Mp_roof','Mp.roof','structural','structural','roof'),
    ('Mp_wall','Mp.wall','structural','structural','wall'),
]
for d in reg:
    ws6.append(d)
    for col_idx in range(1, len(d)+1):
        cell = ws6.cell(row=ws6.max_row, column=col_idx)
        cell.font = df; cell.fill = fills.get('locked', PatternFill()); cell.border = b; cell.alignment = c

ws6.column_dimensions['A'].width = 18
ws6.column_dimensions['B'].width = 14
ws6.column_dimensions['C'].width = 16
ws6.column_dimensions['D'].width = 24
ws6.column_dimensions['E'].width = 10

output_path = '/home/z/my-project/download/steps2-8-unified-model.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
